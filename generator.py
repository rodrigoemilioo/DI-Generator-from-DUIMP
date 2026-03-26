"""
generator.py
------------
Gera o arquivo Excel final no formato semelhante à antiga DI,
com layout profissional, cabeçalho, tabela de adições e totais.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# PALETA DE CORES (estilo Receita Federal / fiscal)
# ─────────────────────────────────────────────────────────────────────────────
C_AZUL_RFB   = "1F3864"   # Azul escuro institucional
C_AZUL_MEDIO = "2E75B6"   # Azul médio
C_AZUL_CLARO = "BDD7EE"   # Azul claro (linha alternada)
C_CINZA      = "F2F2F2"   # Cinza claro
C_AMARELO    = "FFF2CC"   # Amarelo aviso
C_BRANCO     = "FFFFFF"
C_VERDE      = "E2EFDA"   # Verde claro totais
C_VERMELHO   = "C00000"

FONT_BASE = "Arial"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(name=FONT_BASE, bold=bold, color=color, size=size, italic=italic)


def _border(style="thin") -> Border:
    s = Side(border_style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _currency_fmt(ws, cell):
    cell.number_format = '#,##0.00'


# ─────────────────────────────────────────────────────────────────────────────
# GERADOR PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(data: dict, output_path: str) -> str:
    """
    Recebe o dict processado (processor.process_data) e gera o Excel.
    Retorna o caminho do arquivo gerado.
    """
    wb = Workbook()

    # ── Aba 1: Extrato DI ──
    ws1 = wb.active
    ws1.title = "Extrato DI"
    _build_di_sheet(ws1, data)

    # ── Aba 2: Adições Detalhadas ──
    ws2 = wb.create_sheet("Adições Detalhadas")
    _build_adicoes_sheet(ws2, data)

    # ── Aba 3: Resumo Tributário ──
    ws3 = wb.create_sheet("Resumo Tributário")
    _build_resumo_sheet(ws3, data)

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# ABA 1 – EXTRATO DI
# ─────────────────────────────────────────────────────────────────────────────

def _build_di_sheet(ws, data: dict):
    header  = data.get("header", {})
    adicoes = data.get("adicoes", [])
    totais  = data.get("totais", {})
    taxa    = data.get("taxa_cambio_usada", 1.0)

    # ── Larguras de colunas ──
    col_widths = [6, 16, 55, 14, 12, 12, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Título principal ──
    ws.merge_cells(f"A{row}:L{row}")
    c = ws.cell(row=row, column=1,
                value="SECRETARIA DA RECEITA FEDERAL DO BRASIL – RFB")
    c.font      = _font(bold=True, color=C_BRANCO, size=11)
    c.fill      = _fill(C_AZUL_RFB)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"A{row}:L{row}")
    num_di = header.get("numero_duimp", "")
    c = ws.cell(row=row, column=1,
                value=f"EXTRATO DA DECLARAÇÃO DE IMPORTAÇÃO – {num_di}")
    c.font      = _font(bold=True, color=C_BRANCO, size=11)
    c.fill      = _fill(C_AZUL_RFB)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(f"A{row}:L{row}")
    c = ws.cell(row=row, column=1,
                value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                      f"Taxa de câmbio utilizada: R$ {taxa:.4f}/USD")
    c.font      = _font(italic=True, size=9, color="595959")
    c.fill      = _fill(C_CINZA)
    c.alignment = _align("right")
    row += 1

    # ── Bloco Identificação ──
    row = _secao_titulo(ws, row, "IDENTIFICAÇÃO DO PROCESSO", 12)

    campos_id = [
        ("Importador",       header.get("nome_importador", "")),
        ("CNPJ Importador",  header.get("cnpj_importador", "")),
        ("Adquirente",       header.get("nome_adquirente", "")),
        ("CNPJ Adquirente",  header.get("cnpj_adquirente", "")),
        ("Ref. Processo",    header.get("ref_capital", "")),
        ("Ref. Cliente",     header.get("ref_cliente", "")),
        ("Fatura",           header.get("fatura", "")),
        ("Conhecimento",     header.get("conhecimento", "")),
        ("Container",        header.get("container", "")),
        ("Navio",            header.get("navio", "")),
        ("Data Embarque",    header.get("data_embarque", "")),
        ("Data Chegada",     header.get("data_chegada", "")),
        ("País Procedência", header.get("pais_procedencia", "")),
        ("Canal",            header.get("canal", "")),
    ]

    # Renderiza em 2 colunas de pares (label | valor) por linha
    pairs_per_row = 2
    for i in range(0, len(campos_id), pairs_per_row):
        chunk = campos_id[i:i + pairs_per_row]
        # Determina colunas: par 0 → cols A-F, par 1 → cols G-L
        col_starts = [1, 7]
        for j, (label, value) in enumerate(chunk):
            cs = col_starts[j]
            # Label: ocupa 2 colunas
            ws.merge_cells(start_row=row, start_column=cs,
                           end_row=row, end_column=cs + 1)
            lc = ws.cell(row=row, column=cs, value=label)
            lc.font = _font(bold=True, size=9)
            lc.fill = _fill(C_CINZA)
            lc.alignment = _align("right")
            lc.border = _border()
            # Value: ocupa 4 colunas
            ws.merge_cells(start_row=row, start_column=cs + 2,
                           end_row=row, end_column=cs + 5)
            vc = ws.cell(row=row, column=cs + 2, value=value)
            vc.font = _font(size=9)
            vc.alignment = _align()
            vc.border = _border()
        row += 1

    row += 1

    # ── Bloco Valores do Processo ──
    row = _secao_titulo(ws, row, "RESUMO DE VALORES DO PROCESSO", 12)

    val_campos = [
        ("Valor FOB (R$)",       header.get("valor_fob_brl", 0)),
        ("Valor Frete (R$)",     header.get("valor_frete_brl", 0)),
        ("Valor Aduaneiro (R$)", header.get("valor_aduaneiro", 0) or totais.get("valor_aduaneiro", 0)),
    ]
    for label, val in val_campos:
        ws.merge_cells(f"A{row}:C{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True, size=9)
        lc.fill = _fill(C_CINZA)
        lc.alignment = _align("right")
        lc.border = _border()

        ws.merge_cells(f"D{row}:F{row}")
        vc = ws.cell(row=row, column=4, value=val)
        vc.font = _font(size=9)
        vc.number_format = '#,##0.00'
        vc.alignment = _align("right")
        vc.border = _border()
        row += 1

    row += 1

    # ── Tabela de Adições ──
    row = _secao_titulo(ws, row, f"ADIÇÕES ({len(adicoes)} itens)", 12)
    row = _tabela_adicoes(ws, row, adicoes, totais)


def _secao_titulo(ws, row: int, titulo: str, ncols: int) -> int:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=f"  {titulo}")
    c.font      = _font(bold=True, color=C_BRANCO, size=10)
    c.fill      = _fill(C_AZUL_MEDIO)
    c.alignment = _align()
    c.border    = _border()
    ws.row_dimensions[row].height = 16
    return row + 1


def _tabela_adicoes(ws, row: int, adicoes: list, totais: dict) -> int:
    # ── Cabeçalho da tabela ──
    headers = [
        ("Adição", "center"), ("NCM", "center"), ("Descrição do Produto", "left"),
        ("Qtde", "center"), ("Peso Líq.(kg)", "right"), ("Valor Aduan.(R$)", "right"),
        ("II %", "center"), ("II (R$)", "right"),
        ("IPI %", "center"), ("IPI (R$)", "right"),
        ("PIS (R$)", "right"), ("COFINS (R$)", "right"),
        ("ICMS %", "center"), ("ICMS (R$)", "right"),
        ("Total Trib.(R$)", "right"),
    ]

    col_widths_tab = [7, 14, 52, 8, 12, 16, 6, 12, 6, 12, 12, 12, 6, 12, 14]
    for i, w in enumerate(col_widths_tab, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, (h, _) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _font(bold=True, color=C_BRANCO, size=9)
        c.fill      = _fill(C_AZUL_RFB)
        c.alignment = _align("center", wrap=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 28
    row += 1

    # ── Linhas de adição ──
    for idx, ad in enumerate(adicoes):
        bg = C_AZUL_CLARO if idx % 2 == 0 else C_BRANCO

        valores = [
            ad.get("numero_adicao", ""),
            ad.get("ncm", ""),
            _descricao_curta(ad),
            ad.get("quantidade", 0),
            ad.get("peso_liquido", 0),
            ad.get("valor_aduaneiro", 0),
            f"{ad.get('ii_aliquota', 0):.1f}%",
            ad.get("ii_valor", 0),
            f"{ad.get('ipi_aliquota', 0):.1f}%",
            ad.get("ipi_valor", 0),
            ad.get("pis_valor", 0),
            ad.get("cofins_valor", 0),
            f"{ad.get('icms_aliquota', 0):.1f}%",
            ad.get("icms_valor", 0),
            ad.get("total_tributos", 0),
        ]

        aligns = ["center", "center", "left", "right", "right", "right",
                  "center", "right", "center", "right", "right", "right",
                  "center", "right", "right"]

        for col_i, (val, (_, _), aln) in enumerate(
                zip(valores, headers, aligns), start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.font      = _font(size=9)
            c.fill      = _fill(bg)
            c.alignment = _align(aln, wrap=True)
            c.border    = _border()
            if isinstance(val, float) and col_i not in (7, 9, 13):
                c.number_format = '#,##0.00'

        ws.row_dimensions[row].height = 30
        row += 1

    # ── Linha de totais ──
    row = _linha_totais(ws, row, totais, len(headers))
    return row + 2


def _linha_totais(ws, row: int, totais: dict, ncols: int) -> int:
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="TOTAIS")
    c.font      = _font(bold=True, color=C_BRANCO, size=10)
    c.fill      = _fill(C_AZUL_RFB)
    c.alignment = _align("center")
    c.border    = _border()

    valores_totais = {
        6:  totais.get("valor_aduaneiro", 0),
        7:  "",
        8:  totais.get("ii", 0),
        9:  "",
        10: totais.get("ipi", 0),
        11: totais.get("pis", 0),
        12: totais.get("cofins", 0),
        13: "",
        14: totais.get("icms", 0),
        15: totais.get("total_tributos", 0),
    }

    for col_i in range(6, ncols + 1):
        val = valores_totais.get(col_i, "")
        c = ws.cell(row=row, column=col_i, value=val)
        c.font      = _font(bold=True, size=9)
        c.fill      = _fill(C_VERDE)
        c.alignment = _align("right")
        c.border    = _border()
        if isinstance(val, float):
            c.number_format = '#,##0.00'

    ws.row_dimensions[row].height = 18
    return row


def _descricao_curta(ad: dict) -> str:
    """Retorna detalhamento se existir, senão descrição."""
    det = ad.get("detalhamento", "").strip()
    desc = ad.get("descricao", "").strip()
    return det if det else desc


# ─────────────────────────────────────────────────────────────────────────────
# ABA 2 – ADIÇÕES DETALHADAS
# ─────────────────────────────────────────────────────────────────────────────

def _build_adicoes_sheet(ws, data: dict):
    adicoes = data.get("adicoes", [])
    header  = data.get("header", {})

    # Título
    ws.merge_cells("A1:R1")
    c = ws.cell(row=1, column=1,
                value=f"ADIÇÕES DETALHADAS – {header.get('numero_duimp', '')}")
    c.font = _font(bold=True, color=C_BRANCO, size=11)
    c.fill = _fill(C_AZUL_RFB)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 20

    headers_det = [
        "Adição", "NCM", "Cód. Produto", "Descrição", "Detalhamento",
        "País Origem", "Fabricante", "Qtde", "Unidade", "Peso Líq.(kg)",
        "Valor Unit.(USD)", "Valor Total(USD)", "Valor Aduan.(R$)",
        "II %", "II (R$)", "IPI %", "IPI (R$)",
        "PIS %", "PIS (R$)", "COFINS %", "COFINS (R$)",
        "ICMS %", "ICMS (R$)", "Total Trib.(R$)",
        "Aplicação", "Condição",
    ]

    col_w = [7, 14, 12, 45, 55, 20, 35, 8, 8, 12,
             14, 14, 14, 6, 12, 6, 12, 6, 12, 7, 12, 6, 12, 14, 12, 10]
    for i, w in enumerate(col_w, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 2
    for i, h in enumerate(headers_det, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _font(bold=True, color=C_BRANCO, size=9)
        c.fill      = _fill(C_AZUL_MEDIO)
        c.alignment = _align("center", wrap=True)
        c.border    = _border()
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, ad in enumerate(adicoes):
        bg = C_AZUL_CLARO if idx % 2 == 0 else C_BRANCO
        vals = [
            ad.get("numero_adicao", ""),
            ad.get("ncm", ""),
            "",  # código produto (pode vir do XML futuro)
            ad.get("descricao", ""),
            ad.get("detalhamento", ""),
            ad.get("pais_origem", ""),
            ad.get("fabricante", ""),
            ad.get("quantidade", 0),
            ad.get("unidade", "UN"),
            ad.get("peso_liquido", 0),
            ad.get("valor_unitario", 0),
            ad.get("valor_total_usd", 0),
            ad.get("valor_aduaneiro", 0),
            f"{ad.get('ii_aliquota', 0):.1f}%",
            ad.get("ii_valor", 0),
            f"{ad.get('ipi_aliquota', 0):.1f}%",
            ad.get("ipi_valor", 0),
            f"{ad.get('pis_aliquota', 0):.2f}%",
            ad.get("pis_valor", 0),
            f"{ad.get('cofins_aliquota', 0):.2f}%",
            ad.get("cofins_valor", 0),
            f"{ad.get('icms_aliquota', 0):.1f}%",
            ad.get("icms_valor", 0),
            ad.get("total_tributos", 0),
            ad.get("aplicacao", ""),
            ad.get("condicao", ""),
        ]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.font      = _font(size=8)
            c.fill      = _fill(bg)
            c.alignment = _align("left" if col_i in (4, 5, 6, 7) else "right"
                                 if isinstance(val, float) else "center",
                                 wrap=True)
            c.border    = _border()
            if isinstance(val, float):
                c.number_format = '#,##0.00'
        ws.row_dimensions[row].height = 35
        row += 1

    # Freeze panes
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# ABA 3 – RESUMO TRIBUTÁRIO
# ─────────────────────────────────────────────────────────────────────────────

def _build_resumo_sheet(ws, data: dict):
    totais  = data.get("totais", {})
    header  = data.get("header", {})
    adicoes = data.get("adicoes", [])
    taxa    = data.get("taxa_cambio_usada", 1.0)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="RESUMO TRIBUTÁRIO DO PROCESSO")
    c.font = _font(bold=True, color=C_BRANCO, size=12)
    c.fill = _fill(C_AZUL_RFB)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 22
    row += 2

    # Info geral
    info = [
        ("Número DUIMP",         header.get("numero_duimp", "")),
        ("Importador",           header.get("nome_importador", "")),
        ("Adquirente",           header.get("nome_adquirente", "")),
        ("Total de Adições",     str(len(adicoes))),
        ("Taxa de Câmbio (R$/USD)", f"R$ {taxa:.4f}"),
        ("Data de Geração",      datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    for label, val in info:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=True, size=10)
        lc.fill = _fill(C_CINZA)
        lc.border = _border()
        lc.alignment = _align("right")

        ws.merge_cells(f"B{row}:D{row}")
        vc = ws.cell(row=row, column=2, value=val)
        vc.font = _font(size=10)
        vc.border = _border()
        vc.alignment = _align()
        row += 1

    row += 1

    # Tabela de tributos
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="  CONSOLIDADO DE TRIBUTOS")
    c.font = _font(bold=True, color=C_BRANCO, size=10)
    c.fill = _fill(C_AZUL_MEDIO)
    c.alignment = _align()
    row += 1

    cab_headers = ["Tributo", "Valor (R$)", "% s/ Valor Aduaneiro", "Observação"]
    for i, h in enumerate(cab_headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _font(bold=True, color=C_BRANCO, size=10)
        c.fill = _fill(C_AZUL_RFB)
        c.alignment = _align("center")
        c.border = _border()
    row += 1

    va = totais.get("valor_aduaneiro", 1) or 1

    tributos_resumo = [
        ("Valor Aduaneiro", totais.get("valor_aduaneiro", 0), None, "Base de cálculo geral"),
        ("II – Imposto de Importação", totais.get("ii", 0),
         round(totais.get("ii", 0) / va * 100, 2), "Alíquota varia por NCM"),
        ("IPI – Imp. s/ Produtos Industrializados", totais.get("ipi", 0),
         round(totais.get("ipi", 0) / va * 100, 2), "Base = VA + II"),
        ("PIS-Importação", totais.get("pis", 0),
         round(totais.get("pis", 0) / va * 100, 2), "Alíq. 2,10%"),
        ("COFINS-Importação", totais.get("cofins", 0),
         round(totais.get("cofins", 0) / va * 100, 2), "Alíq. 9,65%"),
        ("ICMS-Importação (estimado)", totais.get("icms", 0),
         round(totais.get("icms", 0) / va * 100, 2), "Alíq. base 18% (SC)"),
        ("TOTAL DE TRIBUTOS", totais.get("total_tributos", 0),
         round(totais.get("total_tributos", 0) / va * 100, 2),
         "Soma II + IPI + PIS + COFINS + ICMS"),
    ]

    for i, (nome, valor, pct, obs) in enumerate(tributos_resumo):
        is_total = nome.startswith("TOTAL")
        bg = C_VERDE if is_total else (C_CINZA if i == 0 else
                                       C_AZUL_CLARO if i % 2 == 0 else C_BRANCO)

        c1 = ws.cell(row=row, column=1, value=nome)
        c1.font = _font(bold=is_total, size=10)
        c1.fill = _fill(bg)
        c1.alignment = _align()
        c1.border = _border()

        c2 = ws.cell(row=row, column=2, value=valor)
        c2.font = _font(bold=is_total, size=10)
        c2.fill = _fill(bg)
        c2.alignment = _align("right")
        c2.number_format = 'R$ #,##0.00'
        c2.border = _border()

        c3_val = f"{pct:.2f}%" if pct is not None else "-"
        c3 = ws.cell(row=row, column=3, value=c3_val)
        c3.font = _font(size=10)
        c3.fill = _fill(bg)
        c3.alignment = _align("center")
        c3.border = _border()

        c4 = ws.cell(row=row, column=4, value=obs)
        c4.font = _font(size=9, italic=True)
        c4.fill = _fill(bg)
        c4.alignment = _align()
        c4.border = _border()

        row += 1

    # Nota de rodapé
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    nota = ("NOTA: Os valores de ICMS são estimados com base na alíquota de 18% (SC). "
            "Os impostos calculados automaticamente são aproximações; "
            "confirme com o extrato oficial da DUIMP.")
    c = ws.cell(row=row, column=1, value=nota)
    c.font = _font(size=8, italic=True, color="595959")
    c.fill = _fill(C_AMARELO)
    c.alignment = _align(wrap=True)
    ws.row_dimensions[row].height = 30
